"""Workbook generation tests: synthetic PDB + injected pool, read back with openpyxl."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from athc.pdbtoexcel import Config, PdbWorkbookCreator
from athc.pdbtoexcel.config import default_category_order
from athc.pdbtoexcel.pdb import PDB, PLAY_DATA
from athc.playpool import DefensiveFront
from tests.unit.pdbtoexcel.conftest import (
    make_play_data,
    make_pool,
    make_record,
    make_tendency,
    write_pdb,
)

RUN = PLAY_DATA.PLAY_TYPE.RUN
PASS = PLAY_DATA.PLAY_TYPE.PASS
DEFENSE = PLAY_DATA.PLAY_TYPE.DEFENSE


def _config(**over: object) -> Config:
    base: dict[str, object] = {
        "play_path": "",
        "playpool_rules": None,
        "calculate_total_stats": True,
        "calculate_percentages": True,
        "include_category_worksheets": False,
        "exclude_sacks_from_pass_attempts": True,
    }
    base.update(over)
    return Config(**base)  # type: ignore[arg-type]


def _build(
    tmp_path: Path,
    plays,
    records,
    *,
    config=None,
    calculate_totals=False,
    tendencies=(),
    perform_calcs=True,
):
    pdb = PDB(str(write_pdb(tmp_path / "in.pdb", plays=plays, tendencies=tendencies)))
    out = tmp_path / "out.xlsx"
    creator = PdbWorkbookCreator(
        config or _config(), default_category_order(), make_pool(records), pdb
    )
    creator.create_workbook(str(out), perform_calcs, calculate_totals)
    return openpyxl.load_workbook(out)


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def test_base_sheets_and_run_row(tmp_path: Path) -> None:
    play = make_play_data(
        RUN,
        "Bears",
        "RUN1",
        play_count=10,
        total_yards=46,
        fumbles=1,
        touchdowns_offense=2,
    )
    wb = _build(
        tmp_path, [play], [make_record("RUN1", play_category=0x01, user_category=0x09)]
    )  # Run Middle
    assert {"Options", "Run Plays", "Pass Plays", "Def Plays", "Tendencies"} <= set(
        wb.sheetnames
    )
    header, data = _rows(wb["Run Plays"])[0], _rows(wb["Run Plays"])[1]
    assert header[:6] == ("Team", "Category", "Slot 1", "Slot 2", "Play", "Type")
    assert (data[0], data[1], data[4]) == ("Bears", "Run Middle", "RUN1")
    assert not data[5]  # empty Type (no qb_draw) — openpyxl reads "" as None
    assert (data[6], data[7], data[8]) == (10, 46, 4.6)  # rushes, yards, avg


def test_pass_row_screen_and_stats(tmp_path: Path) -> None:
    play = make_play_data(
        PASS, "Jets", "PASS1", play_count=10, completions=6, sacks=2, total_yards=70
    )
    record = make_record(
        "PASS1", play_category=0x01, user_category=0x03, screen=True
    )  # Pass Short Right
    data = _rows(_build(tmp_path, [play], [record])["Pass Plays"])[1]
    assert (data[1], data[4], data[5]) == ("Pass Short Right", "PASS1", "Screen")
    assert (data[6], data[7]) == (6, 8)  # comp, att (10 - 2 sacks)


def test_defense_row_front_type(tmp_path: Path) -> None:
    play = make_play_data(
        DEFENSE,
        "Vikes",
        "DEF1",
        run_plays_against=5,
        pass_plays_against=5,
        rush_yards_allowed=10,
    )
    record = make_record(
        "DEF1",
        play_category=0x00,
        user_category=0x00,
        defensive_front=DefensiveFront.THREE_FOUR,
    )
    data = _rows(_build(tmp_path, [play], [record])["Def Plays"])[1]
    assert (data[1], data[4], data[5]) == ("Run Right", "DEF1", "3-4")
    assert data[6] == 10  # total calls (run + pass against)


def test_qb_draw_type(tmp_path: Path) -> None:
    play = make_play_data(RUN, "Bears", "QBD", play_count=3, total_yards=9)
    record = make_record("QBD", play_category=0x01, user_category=0x09, qb_draw=True)
    assert _rows(_build(tmp_path, [play], [record])["Run Plays"])[1][5] == "QB draw"


def test_skip_calcs_omits_percent_columns(tmp_path: Path) -> None:
    play = make_play_data(RUN, "Bears", "RUN1", play_count=10, total_yards=46)
    record = make_record("RUN1", play_category=0x01, user_category=0x09)
    with_calcs = _rows(_build(tmp_path, [play], [record])["Run Plays"])[0]
    without = _rows(
        _build(tmp_path, [play], [record], perform_calcs=False)["Run Plays"]
    )[0]
    assert "Fumble %" in with_calcs and "Fumble %" not in without


def test_totals_adds_total_stats_team(tmp_path: Path) -> None:
    plays = [
        make_play_data(RUN, "Bears", "RUN1", play_count=10, total_yards=40),
        make_play_data(RUN, "Jets", "RUN1", play_count=5, total_yards=30),
    ]
    record = make_record("RUN1", play_category=0x01, user_category=0x09)
    teams = [
        row[0]
        for row in _rows(
            _build(tmp_path, plays, [record], calculate_totals=True)["Run Plays"]
        )[1:]
    ]
    assert "Total Stats" in teams


def test_category_worksheets_when_enabled(tmp_path: Path) -> None:
    play = make_play_data(RUN, "Bears", "RUN1", play_count=10, total_yards=40)
    record = make_record("RUN1", play_category=0x01, user_category=0x09)
    wb = _build(
        tmp_path, [play], [record], config=_config(include_category_worksheets=True)
    )
    assert "Run Categories" in wb.sheetnames
    assert _rows(wb["Run Categories"])[1][1] == "Run Middle"


def test_special_teams_and_unknown_plays_skipped(tmp_path: Path) -> None:
    plays = [
        make_play_data(
            RUN, "Bears", "STPLAY", play_count=1
        ),  # in pool but special teams
        make_play_data(RUN, "Bears", "GHOST", play_count=1),  # not in pool
    ]
    record = make_record(
        "STPLAY", play_category=0x01, user_category=0x09, special_category=0x02
    )
    assert _rows(_build(tmp_path, plays, [record])["Run Plays"]) == [
        _rows(_build(tmp_path, [], [])["Run Plays"])[0]
    ]


def test_tendencies_written(tmp_path: Path) -> None:
    wb = _build(tmp_path, [], [], tendencies=[make_tendency("Bears")])
    rows = _rows(wb["Tendencies"])
    assert len(rows) == 1 + 16  # header + 4 downs x 4 buckets
    assert rows[1][0] == "Bears"


def test_slot_column_from_gameplan(tmp_path: Path) -> None:
    from athc.fbpro98_gameplan import read_gameplan
    from tests.unit.pdbtoexcel.conftest import DATA

    plan = read_gameplan(
        str(DATA / "offense.pln")
    )  # OR45RL01 is at normal slot 0 -> "1-1"
    play = make_play_data(RUN, "Bears", "OR45RL01", play_count=5, total_yards=20)
    pdb = PDB(str(write_pdb(tmp_path / "in.pdb", plays=[play])))
    pool = make_pool([make_record("OR45RL01", play_category=0x01, user_category=0x09)])
    out = tmp_path / "out.xlsx"
    creator = PdbWorkbookCreator(
        _config(), default_category_order(), pool, pdb, pln_offense=plan
    )
    creator.create_workbook(str(out), True, False)
    data = _rows(openpyxl.load_workbook(out)["Run Plays"])[1]
    assert data[2] == "1-1"  # Slot 1

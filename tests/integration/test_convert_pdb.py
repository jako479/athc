"""Integration tests for `athc convert-pdb`."""

from __future__ import annotations

import ctypes
import logging
import os
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from athc.cli.convert_pdb import convert_pdb
from athc.pdbtoexcel.pdb import PLAY_DATA
from tests.integration.conftest import DATA

PDB = DATA / "2045-2047.pdb"


# ── extension / usage validation (exit 2) ─────────────────────────────────────


def test_requires_both_args(runner) -> None:
    assert runner.invoke(convert_pdb, [str(PDB)]).exit_code == 2


@pytest.mark.parametrize(
    "args",
    [
        ["in.txt", "out.xlsx"],  # bad pdb extension
        [str(PDB), "out.csv"],  # bad output extension
    ],
)
def test_bad_extension_exit_2(runner, tmp_path: Path, args: list[str]) -> None:
    result = runner.invoke(convert_pdb, [*args, "--play-path", str(tmp_path)])
    assert result.exit_code == 2


def test_bad_pln_extension_exit_2(runner, tmp_path: Path) -> None:
    result = runner.invoke(
        convert_pdb, [str(PDB), str(tmp_path / "o.xlsx"), "-o", "plan.txt"]
    )
    assert result.exit_code == 2


# ── runtime errors (exit 1) ───────────────────────────────────────────────────


def test_missing_pdb_exit_1(
    runner, tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    with caplog.at_level(logging.ERROR):
        result = runner.invoke(
            convert_pdb, [str(tmp_path / "nope.pdb"), str(tmp_path / "o.xlsx")]
        )
    assert result.exit_code == 1
    assert "file not found" in caplog.text


def test_play_path_not_a_directory_exit_1(
    runner, tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    not_a_dir = tmp_path / "notdir.txt"
    not_a_dir.write_text("x", encoding="utf-8")
    with caplog.at_level(logging.ERROR):
        result = runner.invoke(
            convert_pdb,
            [str(PDB), str(tmp_path / "o.xlsx"), "--play-path", str(not_a_dir)],
        )
    assert result.exit_code == 1
    assert "play path is not a directory" in caplog.text


def test_invalid_pdb_content_exit_1(
    runner, tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    bad = tmp_path / "bad.pdb"
    bad.write_bytes(bytes([9]) + b"\x00" * ctypes.sizeof(PLAY_DATA))
    with caplog.at_level(logging.ERROR):
        result = runner.invoke(
            convert_pdb,
            [str(bad), str(tmp_path / "o.xlsx"), "--play-path", str(tmp_path)],
        )
    assert result.exit_code == 1


# ── end-to-end (exit 0) ───────────────────────────────────────────────────────


def test_produces_xlsx_with_sheets(runner, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    result = runner.invoke(
        convert_pdb, [str(PDB), str(out), "--play-path", str(tmp_path)]
    )
    assert result.exit_code == 0
    assert out.is_file()
    wb = openpyxl.load_workbook(out)
    assert {"Options", "Run Plays", "Pass Plays", "Def Plays", "Tendencies"} <= set(
        wb.sheetnames
    )
    # Tendencies come straight from the PDB (no pool needed): 23 teams x 16 rows + header.
    assert len(list(wb["Tendencies"].iter_rows(values_only=True))) == 1 + 23 * 16


def test_produces_xlsm(runner, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsm"
    result = runner.invoke(
        convert_pdb, [str(PDB), str(out), "--play-path", str(tmp_path)]
    )
    assert result.exit_code == 0
    assert out.is_file()


@pytest.mark.parametrize("flag", ["--skip-calcs", "--skip-totals"])
def test_skip_flags(runner, tmp_path: Path, flag: str) -> None:
    out = tmp_path / "out.xlsx"
    result = runner.invoke(
        convert_pdb, [str(PDB), str(out), "--play-path", str(tmp_path), flag]
    )
    assert result.exit_code == 0 and out.is_file()


# ── packaging check (real subprocess) ─────────────────────────────────────────


def test_entry_point_subprocess(tmp_path: Path) -> None:
    env = {**os.environ, "ATHC_CONFIG_DIR": str(tmp_path)}
    out = tmp_path / "out.xlsx"
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "athc",
            "convert-pdb",
            str(PDB),
            str(out),
            "--play-path",
            str(tmp_path),
        ],
        capture_output=True,
        text=True,
        env=env,
    )
    assert result.returncode == 0 and out.is_file()
